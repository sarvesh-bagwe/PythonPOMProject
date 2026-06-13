import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

driver = webdriver.Chrome()

driver.get("https://www.saucedemo.com")
driver.maximize_window()
driver.implicitly_wait(5)

driver.find_element(By.NAME, 'user-name').send_keys("standard_user")
driver.find_element(By.NAME, 'password').send_keys("secret_sauce")

driver.find_element(By.CSS_SELECTOR, '#login-button').click()

element = WebDriverWait(driver, 5).until(ec.visibility_of_element_located((By.CLASS_NAME, 'title')))


print(element.text)


wb = openpyxl.load_workbook('./TestData/shopping_list.xlsx').active

# print(wb.max_column, wb.max_row)

req_prod = []
for i in range(2, wb.max_row +1):
    req_prod.append(wb.cell(row=i, column=1).value)


ele = driver.find_elements(By.XPATH, "//div[@class='inventory_item_label']/a")

assert set(req_prod).issubset(set([e.text for e in ele]))


for l in [e.text for e in ele]:
    driver.find_element(By.XPATH, f"//div[@class='inventory_item_label']/a/div[text()='{l}']"
                                  f"/ancestor::div/following-sibling::div/button").click()
    print("hi")



time.sleep(5)




driver.quit()

